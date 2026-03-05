#!/usr/bin/env python3
"""
seed-from-xlsx.py — Merge Excel issuer catalog into state-registry JSON files.

Reads docs/US_Health_Insurance_Issuers_by_Coverage_Area.xlsx and adds any
issuers missing from the state registry, with known MRF URL patterns or a
browser_required fallback.

Usage: python3 clearnetwork/scripts/seed-from-xlsx.py
"""

import json, re
from datetime import datetime, timezone
from pathlib import Path
import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
REGISTRY_DIR = REPO_ROOT / "clearnetwork" / "state-registry"
XLSX = REPO_ROOT / "docs" / "US_Health_Insurance_Issuers_by_Coverage_Area.xlsx"
NOW = datetime.now(timezone.utc).isoformat()

# ── Known MRF URL patterns per parent company ─────────────────────────────────
# key: lowercase fragment that appears in display names from the Excel
# value: (canonical_name, trade_names, mrf_url, index_type, date_pattern, notes)
KNOWN = {
    "unitedhealthcare": ("UnitedHealth Group", ["UnitedHealthcare","UHC","United Healthcare","Optum","Oxford Health Plans"], "https://transparency-in-coverage.uhc.com/api/v1/uhc/blobs/", "uhc_blob_api", None, ""),
    "anthem": ("Anthem / Elevance Health", ["Anthem","Elevance Health","Anthem Blue Cross Blue Shield","Empire BCBS","Wellpoint"], "https://antm-pt-prod-dataz-nogbd-nophi-us-east1.s3.amazonaws.com/anthem/{date}_anthem_index.json.gz", "dated_s3", "YYYY-MM-01", "403 Forbidden — auth/CORS wall"),
    "elevance": ("Anthem / Elevance Health", ["Anthem","Elevance Health","Anthem Blue Cross Blue Shield"], "https://antm-pt-prod-dataz-nogbd-nophi-us-east1.s3.amazonaws.com/anthem/{date}_anthem_index.json.gz", "dated_s3", "YYYY-MM-01", "403 Forbidden — auth/CORS wall"),
    "cigna": ("Cigna Group / Evernorth", ["Cigna","The Cigna Group","Evernorth","Connecticut General"], "https://d25kgz5rikkq4n.cloudfront.net/cost_transparency/mrf/table-of-contents/reporting_month={date}/{date}_cigna-health-life-insurance-company_index.json", "dated_cloudfront", "YYYY-MM", "403 Forbidden — auth/CORS wall"),
    "aetna": ("Aetna / CVS Health", ["Aetna","CVS Health","CVS Caremark"], "https://mrf.aetna.com/", "browser_required", None, "Known browser-required index type"),
    "humana": ("Humana", ["Humana"], "https://developers.humana.com/syntheticdata/healthplan-price-transparency", "browser_required", None, "Known browser-required index type"),
    "kaiser": ("Kaiser Permanente", ["Kaiser Permanente","Kaiser Foundation Health Plan","KP"], "https://healthy.kaiserpermanente.org/content/dam/kporg/final/documents/health-plan-documents/transparency-in-coverage/kaiser-permanente-all-plans-all-tocs.json", "direct_json", None, ""),
    "centene": ("Centene Corporation", ["Centene","Ambetter","WellCare","Health Net","Fidelis Care","Peach State","Sunshine Health","Managed Health Services","Meridian Health Plan","Coordinated Care"], "https://www.centene.com/price-transparency-files.html", "browser_required", None, "Known browser-required index type"),
    "ambetter": ("Centene Corporation", ["Ambetter","Centene"], "https://www.centene.com/price-transparency-files.html", "browser_required", None, "Known browser-required index type"),
    "molina": ("Molina Healthcare", ["Molina Healthcare","Molina"], "https://www.molinahealthcare.com/members/common/en-US/pdf/medicalpolicies/transparency-coverage.html", "browser_required", None, "Known browser-required index type"),
    "oscar": ("Oscar Health", ["Oscar Health","Oscar"], "https://www.hioscar.com/transparency-in-coverage", "browser_required", None, "Known browser-required index type"),
    "highmark": ("Highmark Health", ["Highmark Blue Cross Blue Shield","Highmark Blue Shield","Highmark BCBS"], "https://mrfdata.hmhs.com/", "browser_required", None, "Known browser-required index type"),
    "caresource": ("CareSource", ["CareSource"], "https://www.caresource.com/about-us/for-providers/forms-and-resources/transparency-in-coverage/", "browser_required", None, ""),
    "bright health": ("Bright Health", ["Bright Health","Bright HealthCare"], "https://www.brighthealthplan.com/transparency-in-coverage", "browser_required", None, ""),
    "friday health": ("Friday Health Plans", ["Friday Health Plans"], "https://fridayhealthplans.com/transparency-in-coverage", "browser_required", None, ""),
    "premera": ("Premera Blue Cross", ["Premera","Premera Blue Cross","Premera BCBS"], "https://www.premera.com/visitor/about-premera/transparency-in-coverage/", "browser_required", None, ""),
    "bcbs of alabama": ("Blue Cross Blue Shield of Alabama", ["BCBS of Alabama","BCBSAL"], "https://www.bcbsal.org/web/healthcare/transparency.html", "browser_required", None, ""),
    "blue cross blue shield of alabama": ("Blue Cross Blue Shield of Alabama", ["BCBSAL"], "https://www.bcbsal.org/web/healthcare/transparency.html", "browser_required", None, ""),
    "bcbs of arizona": ("Blue Cross Blue Shield of Arizona", ["BCBSAZ"], "https://www.azblue.com/about-us/transparency-in-coverage", "browser_required", None, ""),
    "blue cross blue shield of arizona": ("Blue Cross Blue Shield of Arizona", ["BCBSAZ"], "https://www.azblue.com/about-us/transparency-in-coverage", "browser_required", None, ""),
    "arkansas blue": ("Arkansas Blue Cross Blue Shield", ["BCBS Arkansas","Arkansas BCBS"], "https://www.arkansasbluecross.com/members/tools-and-resources/transparency-in-coverage", "browser_required", None, ""),
    "carefirst": ("CareFirst BlueCross BlueShield", ["CareFirst BCBS","CareFirst"], "https://www.carefirst.com/mycarefirst/member/transparencyincoverage.page", "browser_required", None, ""),
    "florida blue": ("Florida Blue (BCBSFL)", ["Florida Blue","BCBSFL","GuideWell"], "https://www.floridablue.com/members/tools-resources/transparency-in-coverage", "browser_required", None, ""),
    "blue cross blue shield of florida": ("Florida Blue (BCBSFL)", ["Florida Blue","BCBSFL"], "https://www.floridablue.com/members/tools-resources/transparency-in-coverage", "browser_required", None, ""),
    "hcsc": ("Health Care Service Corporation (HCSC)", ["HCSC","BCBS of Illinois","BCBS of Texas","BCBS of Montana","BCBS of Oklahoma","BCBS of New Mexico"], "https://www.hcsc.com/transparency-in-coverage", "browser_required", None, ""),
    "blue cross blue shield of illinois": ("Health Care Service Corporation (HCSC)", ["BCBS Illinois","HCSC"], "https://www.bcbsil.com/transparency-in-coverage", "browser_required", None, ""),
    "blue cross blue shield of texas": ("Health Care Service Corporation (HCSC)", ["BCBS Texas","HCSC"], "https://www.bcbstx.com/transparency-in-coverage", "browser_required", None, ""),
    "blue cross blue shield of michigan": ("Blue Cross Blue Shield of Michigan", ["BCBSM"], "https://www.bcbsm.com/legal/transparency-in-coverage.html", "browser_required", None, ""),
    "blue cross blue shield of north carolina": ("Blue Cross Blue Shield of North Carolina", ["BCBSNC"], "https://www.bluecrossnc.com/about-us/transparency-coverage", "browser_required", None, ""),
    "blue cross blue shield of north dakota": ("Blue Cross Blue Shield of North Dakota", ["BCBSND"], "https://www.bcbsnd.com/about-us/transparency-in-coverage", "browser_required", None, ""),
    "bluecross blueshield of south carolina": ("BlueCross BlueShield of South Carolina", ["BCBSSC"], "https://www.bluecrosssc.com/members/resources/transparency-in-coverage", "browser_required", None, ""),
    "blue cross blue shield of tennessee": ("Blue Cross Blue Shield of Tennessee", ["BCBST"], "https://www.bcbst.com/transparency-in-coverage", "browser_required", None, ""),
    "blue cross blue shield of wyoming": ("Blue Cross Blue Shield of Wyoming", ["BCBSWY"], "https://www.bcbswy.com/transparency-in-coverage", "browser_required", None, ""),
    "blue cross blue shield of vermont": ("Blue Cross Blue Shield of Vermont", ["BCBSVT"], "https://www.bcbsvt.com/transparency-in-coverage", "browser_required", None, ""),
    "blue cross blue shield of rhode island": ("Blue Cross Blue Shield of Rhode Island", ["BCBSRI"], "https://www.bcbsri.com/transparency-in-coverage", "browser_required", None, ""),
    "louisiana blue": ("Louisiana Blue (BCBSLA)", ["BCBSLA","Louisiana Blue"], "https://www.bcbsla.com/transparency-in-coverage", "browser_required", None, ""),
    "bcbs of kansas": ("Blue Cross Blue Shield of Kansas", ["BCBSKS"], "https://www.bcbsks.com/transparency-in-coverage", "browser_required", None, ""),
    "blue cross blue shield of kansas": ("Blue Cross Blue Shield of Kansas", ["BCBSKS"], "https://www.bcbsks.com/transparency-in-coverage", "browser_required", None, ""),
    "wellmark": ("Wellmark Blue Cross Blue Shield", ["Wellmark","Wellmark BCBS"], "https://www.wellmark.com/transparency-in-coverage", "browser_required", None, ""),
    "regence": ("Regence BlueCross BlueShield", ["Regence","Regence BCBS","Regence BlueShield"], "https://www.regence.com/transparency-in-coverage", "browser_required", None, ""),
    "blue shield of california": ("Blue Shield of California", ["Blue Shield CA","Blue Shield of California"], "https://www.blueshieldca.com/bsca/bsc/public/sites/shared/en/transparency-in-coverage.sp", "browser_required", None, ""),
    "blue cross of idaho": ("Blue Cross of Idaho", ["BCID","Blue Cross Idaho"], "https://www.bcidaho.com/transparency-in-coverage", "browser_required", None, ""),
    "bcbs of massachusetts": ("Blue Cross Blue Shield of Massachusetts", ["BCBSMA"], "https://www.bluecrossma.org/transparency-in-coverage", "browser_required", None, ""),
    "blue cross blue shield of massachusetts": ("Blue Cross Blue Shield of Massachusetts", ["BCBSMA"], "https://www.bluecrossma.org/transparency-in-coverage", "browser_required", None, ""),
    "harvard pilgrim": ("Harvard Pilgrim Health Care / Point32Health", ["Harvard Pilgrim","Point32Health"], "https://www.harvardpilgrim.org/public/transparency-in-coverage", "browser_required", None, ""),
    "tufts health": ("Tufts Health Plan / Point32Health", ["Tufts Health Plan","Point32Health"], "https://www.tuftshealthplan.com/transparency-in-coverage", "browser_required", None, ""),
    "fallon": ("Fallon Health", ["Fallon Health"], "https://www.fallonhealth.org/transparency-in-coverage", "browser_required", None, ""),
    "health new england": ("Health New England", ["HNE","Health New England"], "https://www.healthnewengland.org/transparency-in-coverage", "browser_required", None, ""),
    "blue cross blue shield of minnesota": ("Blue Cross Blue Shield of Minnesota", ["BCBSMN"], "https://www.bluecrossmn.com/transparency-in-coverage", "browser_required", None, ""),
    "healthpartners": ("HealthPartners", ["HealthPartners"], "https://www.healthpartners.com/hp/insurance/transparency-in-coverage/", "browser_required", None, ""),
    "medica": ("Medica", ["Medica"], "https://www.medica.com/transparency-in-coverage", "browser_required", None, ""),
    "ucare": ("UCare", ["UCare"], "https://www.ucare.org/transparency-in-coverage", "browser_required", None, ""),
    "horizon": ("Horizon Blue Cross Blue Shield of New Jersey", ["Horizon BCBS NJ"], "https://www.horizonblue.com/transparency-in-coverage", "browser_required", None, ""),
    "amerihealth": ("AmeriHealth / Independence Health Group", ["AmeriHealth","Independence Health Group"], "https://www.amerihealth.com/transparency-in-coverage", "browser_required", None, ""),
    "excellus": ("Excellus BlueCross BlueShield", ["Excellus BCBS"], "https://www.excellusbcbs.com/transparency-in-coverage", "browser_required", None, ""),
    "emblemhealth": ("EmblemHealth", ["EmblemHealth","HIP","GHI"], "https://www.emblemhealth.com/transparency-in-coverage", "browser_required", None, ""),
    "connecticare": ("ConnectiCare / EmblemHealth", ["ConnectiCare","EmblemHealth"], "https://www.connecticare.com/transparency-in-coverage", "browser_required", None, ""),
    "mvp health": ("MVP Health Care", ["MVP Health Care","MVP"], "https://www.mvphealthcare.com/transparency-in-coverage", "browser_required", None, ""),
    "healthfirst": ("Healthfirst", ["Healthfirst"], "https://healthfirst.org/transparency-in-coverage", "browser_required", None, ""),
    "metroplus": ("MetroPlus Health Plan", ["MetroPlus"], "https://www.metroplus.org/transparency-in-coverage", "browser_required", None, ""),
    "fidelis care": ("Fidelis Care (Centene)", ["Fidelis Care","Centene"], "https://www.fideliscare.org/transparency-in-coverage", "browser_required", None, ""),
    "independence blue cross": ("Independence Blue Cross", ["IBX","Independence Blue Cross"], "https://www.ibx.com/transparency-in-coverage", "browser_required", None, "IBX EIN-search API"),
    "upmc health plan": ("UPMC Health Plan", ["UPMC","UPMC Health Plan","UPMC for You"], "https://content.upmchp.com/publicweb/table-of-contents/{date}_UPMC-Health-Plan_index.json", "dated_azure", "YYYY-MM-01", ""),
    "priority health": ("Priority Health", ["Priority Health"], "https://www.priorityhealth.com/transparency-in-coverage", "browser_required", None, ""),
    "mclaren": ("McLaren Health Plan", ["McLaren Health Plan","McLaren"], "https://www.mclarenhealthplan.org/transparency-in-coverage", "browser_required", None, ""),
    "quartz": ("Quartz Health Plan", ["Quartz Health Plan","Quartz"], "https://www.quartzbenefits.com/transparency-in-coverage", "browser_required", None, ""),
    "selecthealth": ("SelectHealth", ["SelectHealth","Intermountain Health"], "https://selecthealth.org/transparency-in-coverage", "browser_required", None, ""),
    "mountain health": ("Mountain Health CO-OP", ["Mountain Health CO-OP"], "https://www.mountainhealthco-op.com/transparency-in-coverage", "browser_required", None, ""),
    "community health choice": ("Community Health Choice", ["Community Health Choice"], "https://www.communityhealthchoice.org/transparency-in-coverage", "browser_required", None, ""),
    "la care": ("LA Care Health Plan", ["LA Care","LA Care Health Plan"], "https://www.lacare.org/transparency-in-coverage", "browser_required", None, ""),
    "sharp health": ("Sharp Health Plan", ["Sharp Health Plan"], "https://www.sharphealthplan.com/members/transparency-in-coverage", "browser_required", None, ""),
    "western health advantage": ("Western Health Advantage", ["Western Health Advantage","WHA"], "https://www.westernhealth.com/transparency-in-coverage", "browser_required", None, ""),
    "valley health plan": ("Valley Health Plan", ["Valley Health Plan"], "https://www.valleyhealthplan.org/transparency-in-coverage", "browser_required", None, ""),
    "chinese community": ("Chinese Community Health Plan", ["CCHP","Chinese Community Health Plan"], "https://www.cchphmo.com/transparency-in-coverage", "browser_required", None, ""),
    "neighborhood health plan": ("Neighborhood Health Plan of Rhode Island", ["NHPRI"], "https://www.nhpri.org/transparency-in-coverage", "browser_required", None, ""),
    "community health plan of wa": ("Community Health Plan of Washington", ["CHPW"], "https://www.chpw.org/transparency-in-coverage", "browser_required", None, ""),
}


def match(name):
    nl = name.lower()
    for key, val in KNOWN.items():
        if key in nl:
            return val
    return None


def existing_coverage(registry):
    names = set()
    for e in registry:
        names.add(e["insurer_name"].lower())
        for t in e.get("trade_names", []):
            names.add(t.lower())
    return names


def covered(display_name, existing):
    if display_name.lower() in existing:
        return True
    m = match(display_name)
    if m:
        cname, trades, *_ = m
        if cname.lower() in existing:
            return True
        if any(t.lower() in existing for t in trades):
            return True
    return False


def make_entry(ins_name, trade_names, state, mrf_url, index_type, date_pattern, notes, hios_id):
    e = {
        "insurer_name": ins_name,
        "trade_names": trade_names,
        "state": state,
        "mrf_url": mrf_url,
        "index_type": index_type,
        "date_pattern": date_pattern,
        "cms_source": True,
        "notes": notes,
        "seeded_at": NOW,
    }
    if hios_id:
        e["hios_issuer_id"] = str(hios_id)
    if index_type == "browser_required":
        e["accessibility"] = "browser_required"
        e["transparency_score"] = 0
        e["digital_debt_score"] = 30
    return e


def read_excel():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    issuers = {}  # state -> list of (display_name, hios_id)

    ws = wb["Marketplace Issuers by State"]
    first = True
    for row in ws.iter_rows(values_only=True):
        if first: first = False; continue
        if not row[0]: continue
        state, name, hios_id = str(row[0]).strip(), str(row[1]).strip(), row[2]
        issuers.setdefault(state, []).append((name, hios_id))

    ws = wb["SBE Issuers"]
    first = True
    for row in ws.iter_rows(values_only=True):
        if first: first = False; continue
        if not row[0]: continue
        state, name = str(row[0]).strip(), str(row[2]).strip()
        issuers.setdefault(state, []).append((name, None))

    return issuers


def main():
    print("Reading Excel issuer catalog...")
    excel = read_excel()

    total_added = 0
    updated_states = []

    for state in sorted(excel.keys()):
        reg_path = REGISTRY_DIR / f"{state}.json"
        registry = json.loads(reg_path.read_text()) if reg_path.exists() else []
        existing = existing_coverage(registry)
        new_entries = []
        added_parents = set()  # prevent duplicate parent per state

        for display_name, hios_id in excel[state]:
            if covered(display_name, existing):
                continue

            m = match(display_name)
            if m:
                ins_name, trade_names, mrf_url, index_type, date_pattern, notes = m
                dedup = ins_name.lower()
            else:
                ins_name = display_name
                trade_names = [display_name]
                slug = re.sub(r"[^a-z0-9]", "", display_name.lower())
                mrf_url = f"https://www.{slug}.com/transparency-in-coverage"
                index_type = "browser_required"
                date_pattern = None
                notes = "Regional carrier — MRF URL needs manual verification"
                dedup = ins_name.lower()

            if dedup in added_parents:
                continue
            # Also skip if the canonical name is already in the registry
            if ins_name.lower() in existing or any(t.lower() in existing for t in trade_names):
                existing.update([ins_name.lower()] + [t.lower() for t in trade_names])
                continue

            added_parents.add(dedup)
            existing.update([ins_name.lower()] + [t.lower() for t in trade_names])
            entry = make_entry(ins_name, trade_names, state, mrf_url, index_type, date_pattern, notes, hios_id)
            new_entries.append(entry)
            print(f"  [{state}] + {ins_name}")

        if new_entries:
            registry.extend(new_entries)
            reg_path.write_text(json.dumps(registry, indent=2))
            total_added += len(new_entries)
            updated_states.append(state)

    print(f"\nDone. Added {total_added} new issuers across {len(updated_states)} states: {', '.join(updated_states)}")


if __name__ == "__main__":
    main()
